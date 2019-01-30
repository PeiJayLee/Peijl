from openpyxl import Workbook
import csv

def create_your_workbook(name):
    rows = ['Date', 'ID', 'Platform', 'Volume', 'Rate', 'Close', 'Return', 'Days', 'Changes']
    sum=0
    wb= Workbook()
    ws = wb.active
    ws.title = "new sheet"
    with open(name+'.csv', newline="") as csv_file:
        fr = csv.reader(csv_file)
        for row in fr:
            ws.append(row)
    ws.delete_rows(1)
    for x in ws['A1':'I1']:
        for i in x:
            i.value = rows[sum]
            sum+=1
    for row in ws.iter_rows(min_row=2, min_col=4, max_row='', max_col=4):
        for cell in row:
            try:
                cell.value = int(cell.value)
            except ValueError:
                a = list(cell.value)
                a.remove(',')
                b = ''.join(a)
                cell.value = int(b)
            except TypeError:
                continue
    for row in ws.iter_rows(min_row=2, min_col=5, max_row='', max_col=5):
        for cell in row:
            try:
                cell.value = float(cell.value)
            except TypeError:
                continue
    return wb
